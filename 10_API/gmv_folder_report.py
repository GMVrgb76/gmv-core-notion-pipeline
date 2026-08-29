#!/usr/bin/env python3
"""Create a read-only Markdown inventory of a folder and compare it with a master."""

from __future__ import annotations

import argparse
import hashlib
import mimetypes
import os
import re
import tempfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

TEXT_PREVIEW_LIMIT = 2200
XLSX_PREVIEW_ROWS = 5
TREE_MAX_FILES_PER_DIR = 12
TOOL_VERSION = "0.1.1"

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from PIL import Image
except ImportError:
    Image = None


def human_size(size_bytes: int) -> str:
    units = ["B", "KB", "MB", "GB", "TB"]
    size = float(size_bytes)
    for unit in units:
        if size < 1024 or unit == units[-1]:
            return f"{size:.0f} {unit}" if unit == "B" else f"{size:.2f} {unit}"
        size /= 1024
    return f"{size_bytes} B"


def stable_sha256(
    path: Path, chunk_size: int = 1024 * 1024
) -> tuple[str, os.stat_result]:
    """Hash a regular file and reject a fingerprint if the file changed mid-read."""
    for _attempt in range(2):
        before = path.stat(follow_symlinks=False)
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            while chunk := handle.read(chunk_size):
                digest.update(chunk)
        after = path.stat(follow_symlinks=False)
        before_state = (
            before.st_dev,
            before.st_ino,
            before.st_size,
            before.st_mtime_ns,
        )
        after_state = (after.st_dev, after.st_ino, after.st_size, after.st_mtime_ns)
        if before_state == after_state:
            return digest.hexdigest(), after
    raise OSError("file changed while it was being hashed")


def modified_time(stat_result: os.stat_result) -> str:
    return (
        datetime.fromtimestamp(stat_result.st_mtime)
        .astimezone()
        .isoformat(timespec="seconds")
    )


def safe_text(text: str, limit: int = TEXT_PREVIEW_LIMIT) -> str:
    normalized = " ".join(text.replace("\x00", "").split())
    if len(normalized) > limit:
        return normalized[:limit].rstrip() + " …"
    return normalized


def display_path(path: Path) -> str:
    return safe_text(str(path), limit=10_000)


def inline_code(value: object) -> str:
    text = safe_text(str(value), limit=10_000)
    longest = max(
        (len(match.group(0)) for match in re.finditer(r"`+", text)), default=0
    )
    fence = "`" * max(1, longest + 1)
    padding = " " if text.startswith("`") or text.endswith("`") else ""
    return f"{fence}{padding}{text}{padding}{fence}"


def fenced_block(text: str, language: str = "text") -> list[str]:
    longest = max(
        (len(match.group(0)) for match in re.finditer(r"`+", text)), default=0
    )
    fence = "`" * max(3, longest + 1)
    return [f"{fence}{language}", text, fence]


def heading_text(value: str) -> str:
    text = safe_text(value, limit=10_000)
    return re.sub(r"([\\`*_{}\[\]<>#+.!|])", r"\\\1", text)


def extract_pdf(path: Path) -> dict:
    if PdfReader is None:
        return {"supported": False, "note": "pypdf non installato"}
    try:
        reader = PdfReader(str(path))
        parts = []
        for page in reader.pages[:8]:
            text = page.extract_text() or ""
            if text.strip():
                parts.append(text)
            if sum(len(part) for part in parts) >= TEXT_PREVIEW_LIMIT * 2:
                break
        full_text = "\n".join(parts).strip()
        return {
            "supported": True,
            "pages": len(reader.pages),
            "has_text": bool(full_text),
            "preview": safe_text(full_text) if full_text else "",
            "likely_scanned": not bool(full_text),
        }
    except Exception as exc:  # third-party parsers expose heterogeneous errors
        return {"supported": True, "error": safe_text(str(exc))}


def extract_docx(path: Path) -> dict:
    if Document is None:
        return {"supported": False, "note": "python-docx non installato"}
    try:
        document = Document(str(path))
        paragraphs = [
            item.text.strip() for item in document.paragraphs if item.text.strip()
        ]
        preview = safe_text("\n".join(paragraphs))
        return {"supported": True, "has_text": bool(preview), "preview": preview}
    except Exception as exc:  # third-party parsers expose heterogeneous errors
        return {"supported": True, "error": safe_text(str(exc))}


def extract_text_file(path: Path) -> dict:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        return {
            "supported": True,
            "has_text": bool(text.strip()),
            "preview": safe_text(text),
        }
    except OSError as exc:
        return {"supported": True, "error": safe_text(str(exc))}


def extract_xlsx(path: Path) -> dict:
    if load_workbook is None:
        return {"supported": False, "note": "openpyxl non installato"}
    try:
        workbook = load_workbook(
            filename=str(path), read_only=True, data_only=True, keep_links=False
        )
        sheets = []
        for worksheet in workbook.worksheets:
            rows = []
            for row in worksheet.iter_rows(
                min_row=1, max_row=XLSX_PREVIEW_ROWS + 1, values_only=True
            ):
                rows.append(["" if value is None else str(value) for value in row])
            sheets.append(
                {
                    "title": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "rows": rows,
                }
            )
        workbook.close()
        return {"supported": True, "sheets": sheets}
    except Exception as exc:  # third-party parsers expose heterogeneous errors
        return {"supported": True, "error": safe_text(str(exc))}


def extract_image(path: Path) -> dict:
    if Image is None:
        return {"supported": False, "note": "Pillow non installato"}
    try:
        with Image.open(path) as image:
            return {
                "supported": True,
                "format": image.format,
                "width": image.width,
                "height": image.height,
            }
    except Exception as exc:  # third-party parsers expose heterogeneous errors
        return {"supported": True, "error": safe_text(str(exc))}


def extract_metadata(path: Path) -> dict:
    extension = path.suffix.lower()
    if extension == ".pdf":
        return extract_pdf(path)
    if extension == ".docx":
        return extract_docx(path)
    if extension in {".txt", ".md", ".csv", ".tsv", ".json", ".yaml", ".yml", ".xml"}:
        return extract_text_file(path)
    if extension in {".xlsx", ".xlsm"}:
        return extract_xlsx(path)
    if extension in {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".webp", ".bmp", ".gif"}:
        return extract_image(path)
    return {"supported": False, "note": "formato non estratto"}


def collect_files(
    root: Path, excluded: set[Path] | None = None
) -> tuple[list[Path], list[str], int]:
    """Collect regular files without following symlinks; retain incomplete-scan evidence."""
    files: list[Path] = []
    problems: list[str] = []
    folder_count = 0
    excluded = excluded or set()

    def record_walk_error(error: OSError) -> None:
        target = getattr(error, "filename", None) or root
        problems.append(f"{target}: {error}")

    for current, directory_names, file_names in os.walk(
        root, topdown=True, followlinks=False, onerror=record_walk_error
    ):
        current_path = Path(current)
        directory_names.sort(key=str.casefold)
        file_names.sort(key=str.casefold)

        safe_directories = []
        for name in directory_names:
            path = current_path / name
            try:
                if path.is_symlink():
                    problems.append(
                        f"{path.relative_to(root)}: symlink directory skipped"
                    )
                else:
                    safe_directories.append(name)
                    folder_count += 1
            except OSError as exc:
                problems.append(f"{path.relative_to(root)}: {exc}")
        directory_names[:] = safe_directories

        for name in file_names:
            path = current_path / name
            try:
                if path.is_symlink():
                    problems.append(f"{path.relative_to(root)}: symlink file skipped")
                    continue
                if path.absolute() in excluded:
                    continue
                if path.is_file():
                    files.append(path)
            except OSError as exc:
                problems.append(f"{path.relative_to(root)}: {exc}")

    files.sort(key=lambda item: str(item.relative_to(root)).casefold())
    return files, problems, folder_count


def build_tree(root: Path, excluded: set[Path] | None = None) -> str:
    lines = [display_path(Path(root.name or str(root)))]
    excluded = excluded or set()

    def walk(folder: Path, prefix: str = "") -> None:
        try:
            entries = sorted(os.scandir(folder), key=lambda item: item.name.casefold())
        except OSError as exc:
            lines.append(prefix + f"└── [UNREADABLE: {safe_text(str(exc))}]")
            return

        directories = []
        files = []
        special = []
        for entry in entries:
            entry_path = Path(entry.path)
            try:
                if entry.is_symlink():
                    target = safe_text(os.readlink(entry.path), limit=10_000)
                    special.append(f"{entry.name}@ -> {target} [SKIPPED]")
                elif entry.is_dir(follow_symlinks=False):
                    directories.append(entry)
                elif (
                    entry.is_file(follow_symlinks=False)
                    and entry_path.absolute() not in excluded
                ):
                    files.append(entry)
                elif entry_path.absolute() not in excluded:
                    special.append(f"{entry.name} [SPECIAL FILE SKIPPED]")
            except OSError as exc:
                special.append(f"{entry.name} [UNREADABLE: {safe_text(str(exc))}]")

        visible_files = files[:TREE_MAX_FILES_PER_DIR]
        hidden_count = len(files) - len(visible_files)
        shown: list[tuple[str, object]] = [("directory", item) for item in directories]
        shown.extend(("file", item) for item in visible_files)
        shown.extend(("special", item) for item in special)
        if hidden_count:
            shown.append(("hidden", hidden_count))

        for index, (kind, item) in enumerate(shown):
            last = index == len(shown) - 1
            connector = "└── " if last else "├── "
            child_prefix = prefix + ("    " if last else "│   ")
            if kind == "directory":
                lines.append(
                    prefix + connector + safe_text(item.name, limit=10_000) + "/"
                )
                walk(Path(item.path), child_prefix)
            elif kind == "file":
                lines.append(prefix + connector + safe_text(item.name, limit=10_000))
            elif kind == "special":
                lines.append(prefix + connector + safe_text(str(item), limit=10_000))
            else:
                lines.append(prefix + connector + f"… {item} altri file")

    walk(root)
    return "\n".join(lines)


def index_master(
    master_root: Optional[Path], excluded: set[Path]
) -> tuple[dict[str, list[str]], dict[str, list[str]], list[str], int]:
    by_hash: dict[str, list[str]] = defaultdict(list)
    by_name: dict[str, list[str]] = defaultdict(list)
    if not master_root:
        return by_hash, by_name, [], 0

    files, problems, _folder_count = collect_files(master_root, excluded)
    indexed_count = 0
    for path in files:
        relative = display_path(path.relative_to(master_root))
        by_name[path.name.casefold()].append(relative)
        try:
            digest, _stat_result = stable_sha256(path)
        except OSError as exc:
            problems.append(f"{relative}: hash failed: {exc}")
            continue
        by_hash[digest].append(relative)
        indexed_count += 1
    return by_hash, by_name, problems, indexed_count


def markdown_table_row(values: list[object]) -> str:
    escaped = []
    for value in values:
        text = safe_text(str(value), limit=10_000)
        escaped.append(text.replace("\\", "\\\\").replace("|", r"\|"))
    return "| " + " | ".join(escaped) + " |"


def atomic_write(output: Path, text: str) -> None:
    if not output.parent.is_dir():
        raise OSError(f"output parent is not a directory: {output.parent}")
    temporary_name = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=output.parent,
            prefix=f".{output.name}.",
            delete=False,
        ) as handle:
            temporary_name = handle.name
            handle.write(text)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_name, output)
    finally:
        if temporary_name and os.path.exists(temporary_name):
            os.unlink(temporary_name)


def report_folder(source: Path, master: Optional[Path], output: Path) -> None:
    excluded = {output.absolute()}
    files, source_problems, folder_count = collect_files(source, excluded)
    master_by_hash, master_by_name, master_problems, master_indexed = index_master(
        master, excluded
    )
    master_complete = not master_problems

    records = []
    hash_groups: dict[str, list[str]] = defaultdict(list)
    extension_counts = Counter()
    total_size = 0

    for path in files:
        relative = display_path(path.relative_to(source))
        try:
            digest, stat_result = stable_sha256(path)
        except OSError as exc:
            source_problems.append(f"{relative}: hash/stat failed: {exc}")
            continue

        size = stat_result.st_size
        extension = path.suffix.lower() or "[no extension]"
        total_size += size
        extension_counts[extension] += 1
        hash_groups[digest].append(relative)
        metadata = extract_metadata(path)

        records.append(
            {
                "id": f"F{len(records) + 1:04d}",
                "rel": relative,
                "name": safe_text(path.name, limit=10_000),
                "ext": extension,
                "mime": mimetypes.guess_type(path.name)[0] or "unknown",
                "size": size,
                "modified": modified_time(stat_result),
                "sha256": digest,
                "meta": metadata,
                "exact_master": master_by_hash.get(digest, []),
                "name_master": master_by_name.get(path.name.casefold(), []),
            }
        )

    exact_duplicates = {
        digest: paths for digest, paths in hash_groups.items() if len(paths) > 1
    }
    generated = datetime.now().astimezone().isoformat(timespec="seconds")
    lines = [
        "# GMV Folder Report",
        "",
        f"**Source:** {inline_code(display_path(source))}  ",
        f"**GMV Master:** {inline_code(display_path(master))}  "
        if master
        else "**GMV Master:** not supplied  ",
        f"**Generated:** {generated}  ",
        f"**Files successfully inventoried:** {len(records)}  ",
        f"**Folders discovered:** {folder_count}  ",
        f"**Total size:** {human_size(total_size)}",
        "",
        "## Directory structure",
        "",
        *fenced_block(build_tree(source, excluded)),
        "",
        "## File types",
        "",
    ]

    if extension_counts:
        for extension, count in sorted(
            extension_counts.items(), key=lambda item: (-item[1], item[0])
        ):
            lines.append(f"- {inline_code(extension)}: {count}")
    else:
        lines.append("No regular files inventoried.")

    lines.extend(["", "## Summary index", ""])
    lines.append("| ID | File | Type | Size | Text/Info | GMV Master | Path |")
    lines.append("|---|---|---|---:|---|---|---|")

    for record in records:
        metadata = record["meta"]
        text_info = "NO"
        if metadata.get("preview"):
            text_info = "YES"
        elif record["ext"] in {
            ".jpg",
            ".jpeg",
            ".png",
            ".tif",
            ".tiff",
            ".webp",
            ".bmp",
            ".gif",
        } and metadata.get("supported"):
            text_info = "IMAGE META"
        elif record["ext"] in {".xlsx", ".xlsm"} and metadata.get("supported"):
            text_info = "SHEET META"

        if record["exact_master"]:
            master_status = "EXACT MATCH"
        elif record["name_master"]:
            master_status = "SAME NAME"
        elif master and not master_complete:
            master_status = "UNKNOWN — INDEX INCOMPLETE"
        else:
            master_status = "NOT FOUND" if master else "NOT CHECKED"

        lines.append(
            markdown_table_row(
                [
                    record["id"],
                    record["name"],
                    record["ext"],
                    human_size(record["size"]),
                    text_info,
                    master_status,
                    record["rel"],
                ]
            )
        )

    lines.extend(["", "## Exact duplicates inside source", ""])
    if exact_duplicates:
        for digest, paths in exact_duplicates.items():
            lines.append(f"- {inline_code(digest[:12] + '…')}")
            for path in paths:
                lines.append(f"  - {inline_code(path)}")
    else:
        lines.append("No exact duplicates detected.")

    if master:
        exact_count = sum(bool(record["exact_master"]) for record in records)
        name_count = sum(
            bool(record["name_master"]) and not record["exact_master"]
            for record in records
        )
        unmatched_count = len(records) - exact_count - name_count
        unmatched_label = (
            "Not found" if master_complete else "Unresolved due to incomplete index"
        )
        lines.extend(
            [
                "",
                "## Comparison with GMV_MASTER_SYSTEM",
                "",
                f"- Master files successfully indexed: **{master_indexed}**",
                f"- Master index complete: **{'YES' if master_complete else 'NO'}**",
                f"- Exact file matches: **{exact_count}**",
                f"- Same filename but different content: **{name_count}**",
                f"- {unmatched_label}: **{unmatched_count}**",
            ]
        )

    lines.extend(["", "## Scan evidence", ""])
    if source_problems:
        lines.append(
            f"Source scan complete: **NO** — {len(source_problems)} skipped/error item(s)."
        )
        for problem in source_problems:
            lines.append(f"- {inline_code(problem)}")
    else:
        lines.append("Source scan complete: **YES**")
    if master:
        if master_problems:
            lines.append(
                f"Master index complete: **NO** — {len(master_problems)} skipped/error item(s)."
            )
            for problem in master_problems:
                lines.append(f"- {inline_code(problem)}")
        else:
            lines.append("Master index complete: **YES**")

    lines.extend(["", "## File details", ""])
    for record in records:
        metadata = record["meta"]
        lines.extend(
            [
                f"### {record['id']} — {heading_text(record['name'])}",
                "",
                f"- Path: {inline_code(record['rel'])}",
                f"- Type: {inline_code(record['ext'])} / {inline_code(record['mime'])}",
                f"- Size: {human_size(record['size'])}",
                f"- Modified: {record['modified']}",
                f"- SHA256: {inline_code(record['sha256'])}",
            ]
        )

        if record["exact_master"]:
            lines.append("- GMV Master: **EXACT MATCH**")
            for path in record["exact_master"]:
                lines.append(f"  - {inline_code(path)}")
        elif record["name_master"]:
            lines.append("- GMV Master: **SAME FILENAME FOUND**")
            for path in record["name_master"]:
                lines.append(f"  - {inline_code(path)}")
        elif master and master_complete:
            lines.append("- GMV Master: **NOT FOUND**")
        elif master:
            lines.append("- GMV Master: **UNKNOWN — INDEX INCOMPLETE**")

        if "pages" in metadata:
            lines.append(f"- PDF pages: {metadata['pages']}")
        if metadata.get("likely_scanned"):
            lines.append("- Text extraction: **NO — likely scanned PDF**")
        elif metadata.get("preview"):
            lines.append("- Text extraction: **YES**")
        if metadata.get("format"):
            lines.append(
                f"- Image: {metadata.get('format')} — {metadata.get('width')}×{metadata.get('height')} px"
            )
        if metadata.get("note"):
            lines.append(f"- Extraction note: {safe_text(str(metadata['note']))}")
        if metadata.get("error"):
            lines.append(f"- Extraction error: {inline_code(metadata['error'])}")
        if metadata.get("preview"):
            lines.extend(
                ["", "**Content preview**", "", *fenced_block(metadata["preview"])]
            )
        if metadata.get("sheets"):
            lines.extend(["", "**Workbook preview**", ""])
            for sheet in metadata["sheets"]:
                lines.append(
                    f"- Sheet {inline_code(sheet['title'])} — rows: {sheet['max_row']}, columns: {sheet['max_column']}"
                )
                rows = sheet["rows"]
                if rows:
                    preview = "\n".join(" | ".join(row) for row in rows)
                    lines.extend(["", *fenced_block(preview)])
        lines.append("")

    lines.extend(
        [
            "## Operational note",
            "",
            "This report is descriptive only.",
            "No source or master file was modified, moved, renamed or deleted.",
            "Only the requested report output was written.",
            "",
        ]
    )
    atomic_write(output, "\n".join(lines))


def existing_directory(value: str) -> Path:
    path = Path(value).expanduser().resolve()
    if not path.is_dir():
        raise argparse.ArgumentTypeError(f"not a valid directory: {path}")
    return path


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Produce un report Markdown di una cartella e la confronta con GMV_MASTER_SYSTEM."
    )
    parser.add_argument(
        "--version", action="version", version=f"%(prog)s {TOOL_VERSION}"
    )
    parser.add_argument(
        "source", type=existing_directory, help="Cartella sorgente da analizzare"
    )
    parser.add_argument(
        "--master",
        type=existing_directory,
        default=None,
        help="Percorso locale di GMV_MASTER_SYSTEM",
    )
    parser.add_argument(
        "-o", "--output", type=Path, default=None, help="File Markdown di output"
    )
    return parser.parse_args(argv)


def main(argv: Optional[list[str]] = None) -> int:
    args = parse_args(argv)
    source: Path = args.source
    master: Optional[Path] = args.master
    if master == source:
        raise SystemExit("Errore: source e --master devono essere cartelle diverse")

    default_name = re.sub(r"[^A-Za-z0-9._-]+", "_", source.name).strip("._") or "folder"
    output = (
        args.output.expanduser().resolve()
        if args.output
        else Path.cwd() / f"{default_name}_REPORT.md"
    )
    if output.exists() and output.is_dir():
        raise SystemExit(f"Errore: output è una cartella: {output}")

    try:
        report_folder(source, master, output)
    except OSError as exc:
        raise SystemExit(f"Errore durante la generazione del report: {exc}") from exc
    print(f"Report written: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
